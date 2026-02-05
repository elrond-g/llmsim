from typing import Any

from src.arch.perf_calculator import ModelPerformance
from src.visual.report_base import ReportFormatter


class ExcelReportFormatter(ReportFormatter):
    """Excel output formatter"""

    def format(self, model_perf: ModelPerformance) -> Any:
        """
        Format performance report as Excel workbook

        Args:
            model_perf: Model performance metrics

        Returns:
            openpyxl.Workbook object
        """
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        except ImportError:
            raise ImportError(
                "openpyxl library is required for Excel output."
                "Please run: pip install openpyxl"
            )

        all_rows = self._collect_data(model_perf)

        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Performance Analysis"

        # Set column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["E"].width = 8
        ws.column_dimensions["F"].width = 8
        ws.column_dimensions["G"].width = 8
        ws.column_dimensions["H"].width = 10
        ws.column_dimensions["I"].width = 10
        ws.column_dimensions["J"].width = 10
        ws.column_dimensions["K"].width = 12
        ws.column_dimensions["L"].width = 12
        ws.column_dimensions["M"].width = 12
        ws.column_dimensions["N"].width = 12
        ws.column_dimensions["O"].width = 12
        ws.column_dimensions["P"].width = 10
        ws.column_dimensions["Q"].width = 12

        # Define styles
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=12)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")

        # Add title
        ws["A1"] = (
            f"Performance Analysis Report: {model_perf.model_name} ({model_perf.forward_mode})"
        )
        ws["A1"].font = title_font
        ws.merge_cells("A1:Q1")
        ws["A1"].alignment = center_align

        # Add column headers
        headers = [
            "Operator Name",
            "Type",
            "m",
            "n",
            "k",
            "batch",
            "layers",
            "Input",
            "Output",
            "Weight",
            "Compute(us)",
            "Memory(us)",
            "Transfer(us)",
            "Single Layer Theory Latency(us)",
            "Total Time(ms)",
            "Percent(%)",
            "Weight/Single GPU All Layers",
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        # 添加数据行
        col_keys = [
            "name",
            "type",
            "m",
            "n",
            "k",
            "batch",
            "layers",
            "in_dtype",
            "out_dtype",
            "weight_dtype",
            "compute",
            "memory",
            "transfer",
            "op_time_single_layer",
            "total",
            "percent",
            "op_weight_mem",
        ]

        for row_idx, row_data in enumerate(all_rows, start=4):
            for col_idx, key in enumerate(col_keys, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row_data[key]

                # 格式化值
                if key in ["compute", "memory", "transfer", "total"]:
                    cell.value = round(value, 3)
                    cell.number_format = "0.000"
                    cell.alignment = right_align
                elif key == "op_time_single_layer":
                    cell.value = round(value, 3)
                    cell.number_format = "0.000"
                    cell.alignment = right_align
                elif key == "percent":
                    cell.value = round(value, 2)
                    cell.number_format = "0.00"
                    cell.alignment = right_align
                elif key in ["m", "n", "k", "batch", "layers", "op_weight_mem"]:
                    cell.value = value
                    cell.alignment = right_align
                else:
                    cell.value = str(value)
                    cell.alignment = left_align

                cell.border = border

        # Add statistics row
        stats_row = len(all_rows) + 7
        total_compute_ms = model_perf.total_compute_time / 1000.0
        total_memory_ms = model_perf.total_memory_time / 1000.0
        total_transfer_ms = model_perf.total_transfer_time / 1000.0
        total_ms = model_perf.total_time / 1000.0

        stats_data = [
            ("Compute Time (ms)", total_compute_ms),
            ("Memory Time (ms)", total_memory_ms),
            ("Transfer Time (ms)", total_transfer_ms),
            ("Total Time (ms)", total_ms),
        ]

        for idx, (label, value) in enumerate(stats_data):
            label_cell = ws.cell(row=stats_row + idx, column=1)
            label_cell.value = label
            label_cell.font = Font(bold=True)

            value_cell = ws.cell(row=stats_row + idx, column=2)
            value_cell.value = round(value, 3)
            value_cell.number_format = "0.000"

        # Add performance bottleneck information
        bottleneck = model_perf.get_bottleneck_op()
        if bottleneck:
            bottleneck_row = stats_row + len(stats_data) + 2
            label_cell = ws.cell(row=bottleneck_row, column=1)
            label_cell.value = "Performance Bottleneck"
            label_cell.font = Font(bold=True)

            _, op_name, op_perf = bottleneck
            value_cell = ws.cell(row=bottleneck_row, column=2)
            value_cell.value = f"{op_name} (Total Time: {op_perf.total_time:.3f} ms)"

        # TTFT and Throughput
        ttft = model_perf.get_ttft_or_tpot()
        throughput = model_perf.get_throughput()
        other_data = [
            (
                ("TTFT (ms)", ttft)
                if model_perf.forward_mode == "EXTEND"
                else ("TPOT (ms)", ttft)
            ),
            ("Throughput TPS", throughput),
            (
                "Weight Memory/Single GPU (GB)",
                model_perf.model_total_mem_occupy / (1024**3),
            ),
        ]
        for idx, (label, value) in enumerate(other_data):
            ttft_throughput_row = (
                stats_row + len(stats_data) + 2 + (2 if bottleneck else 0)
            )
            label_cell = ws.cell(row=ttft_throughput_row + idx, column=1)
            label_cell.value = label
            label_cell.font = Font(bold=True)

            value_cell = ws.cell(row=ttft_throughput_row + idx, column=2)
            value_cell.value = round(value, 3)
            value_cell.number_format = "0.000"

        return wb

    def save(self, model_perf: ModelPerformance, output_path: str = None) -> None:
        """
        Save performance report to Excel file

        Args:
            model_perf: Model performance metrics
            output_path: Output Excel file path
        """
        if not output_path:
            output_path = "Performance_Report.xlsx"

        try:
            wb = self.format(model_perf)
            wb.save(output_path)
            print(f"Excel report saved to: {output_path}")
        except Exception as e:
            print(f"Failed to save Excel report: {e}")
